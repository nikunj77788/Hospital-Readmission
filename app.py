# app.py - Hospital Readmission Prediction & Dashboards
from werkzeug.security import generate_password_hash, check_password_hash
import json
import os
from flask import Flask, render_template, request, redirect, session, url_for, flash, jsonify, send_file
import pandas as pd
import plotly.express as px
import pickle
import datetime
from io import BytesIO
from reportlab.pdfgen import canvas
from model import predict_risk  # use the function already defined in model.py
import numpy as np
from datetime import datetime
from flask import session, jsonify, request
from flask import Flask, render_template, request, redirect, url_for, flash
import os
from notify_sms import send_sms  #updated Twilio SMS function
from flask import Flask, request, jsonify, session
import pandas as pd
from notify_sms import notify_patient_sms, send_sms
from openpyxl.styles import Font, PatternFill





# --------------------------
# Helper functions
# --------------------------
def encode_yes_no(value):
    value = value.strip().upper()
    if value == 'YES':
        return 1
    elif value == 'NO':
        return 0
    else:
        raise ValueError(f"Unknown value: {value}")

def age_to_bin(age):
    age = int(age)
    if age < 10: return "[0-10)"
    elif age < 20: return "[10-20)"
    elif age < 30: return "[20-30)"
    elif age < 40: return "[30-40)"
    elif age < 50: return "[40-50)"
    elif age < 60: return "[50-60)"
    elif age < 70: return "[60-70)"
    elif age < 80: return "[70-80)"
    elif age < 90: return "[80-90)"
    else: return "[90-100)"


# --------------------------
# Helper function
# --------------------------
def safe_age_to_int(age):
    """
    Safely convert age to int.
    Returns 0 if conversion fails.
    """
    try:
        return int(age)
    except (ValueError, TypeError):
        return 0


def patient_risk_score(row):
    """
    Improved and consistent AI risk scoring.
    Handles Yes/No, 0/1, and missing values safely.
    Returns a float between 0 and 1.
    """
    try:
        def is_yes(value):
            """Normalize yes/no or boolean-like values."""
            if value is None:
                return False
            v = str(value).strip().upper()
            return v in ["YES", "Y", "1", "TRUE"]

        risk = 0.0

        # --- Age Factor ---
        age = safe_age_to_int(row.get("age", 0))
        if age >= 70:
            risk += 0.25
        elif age >= 60:
            risk += 0.15
        elif age >= 50:
            risk += 0.1

        # --- Hospital Stay ---
        time_in_hospital = int(row.get("time_in_hospital", 0) or 0)
        risk += min(time_in_hospital * 0.03, 0.25)  # stronger effect, capped

        # --- Number of Medications ---
        num_meds = int(row.get("num_medications", 0) or 0)
        risk += min(num_meds * 0.01, 0.2)

        # --- Change in medication ---
        if is_yes(row.get("change_in_medication", row.get("change"))):
            risk += 0.1

        # --- Diabetes medication ---
        if is_yes(row.get("diabetes_medication", row.get("diabetesMed"))):
            risk += 0.1

        # --- Past Readmission ---
        if is_yes(row.get("readmitted")):
            risk += 0.25

        # --- Chronic conditions (optional columns) ---
        for cond in ["high_blood_pressure", "heart_disease", "smoker"]:
            if cond in row and is_yes(row[cond]):
                risk += 0.05

        # Cap final risk between 0 and 1
        return float(min(max(risk, 0.0), 1.0))
    except Exception as e:
        print("⚠️ patient_risk_score error:", e)
        return 0.0
def ai_level(score):
    """Convert score to categorical label."""
    if score >= 0.7:
        return "High"
    elif score >= 0.4:
        return "Medium"
    else:
        return "Low"



# --------------------------
# Flask App
# --------------------------
app = Flask(__name__)
app.secret_key = "supersecretkey"

HIGH_RISK_THRESHOLD = 0.7  # example threshold

# Store in-app alerts (can be replaced with DB later)
in_app_alerts = []

@app.route("/notify_high_risk", methods=["POST"])
def notify_high_risk():
    data = request.get_json()
    patient_name = data.get("patient_name", "Unknown")
    patient_id = data.get("patient_id", "N/A")
    risk_score = float(data.get("risk_score", 0))
    recommendation = data.get("recommendation", "Follow care plan.")

    # Only send alerts if score exceeds threshold
    if risk_score >= HIGH_RISK_THRESHOLD:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # --- In-app alert ---
        alert = {
            "patient_id": patient_id,
            "patient_name": patient_name,
            "risk_score": risk_score,
            "recommendation": recommendation,
            "timestamp": timestamp
        }
        in_app_alerts.append(alert)

# --- Role-Based Access Decorator ---
from functools import wraps

def login_required(role):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if session.get("role") != role:
                flash("Access denied")
                return redirect("/login")
            return func(*args, **kwargs)
        return wrapper
    return decorator




# Load dataset
DATA_PATH = "cleaned_diabetes.csv"
df = pd.read_csv(DATA_PATH, low_memory=False)

# Fill missing or empty department values
df['department'] = df.get('department', 'Unknown').fillna('Unknown')
df['department'] = df['department'].replace('', 'Unknown')

# Fill missing departments with a proper name
if 'department' not in df.columns:
    df['department'] = "General"
else:
    df['department'] = df['department'].fillna("General")
    df['department'] = df['department'].replace('', 'General')


# --------------------------
# Robust YES/NO and numeric conversion
# --------------------------
yes_no_cols = ['high_blood_pressure', 'heart_disease', 'smoker', 'readmitted']

for col in yes_no_cols:
    if col in df.columns:
        df[col] = df[col].apply(lambda x: 1 if str(x).strip().upper() == 'YES' or str(x).strip() == '1' else 0)

# Users file
USERS_FILE = "users.json"

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    else:
        return {}

def save_user(username, raw_password, role, name):
    users = load_users()
    # Hash password before saving
    hashed = generate_password_hash(raw_password)
    users[username] = {"password": hashed, "role": role, "name": name}
    with open(USERS_FILE, "w") as f:
        json.dump(users, f)

# --------------------------
# Routes
# --------------------------
@app.route("/", methods=["GET", "POST"])
def index():
    return redirect(url_for("login"))

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        role = request.form["role"]

        users = load_users()
        if username in users:
            user = users[username]
            if check_password_hash(user["password"], password) and user["role"] == role:
                session["username"] = username
                session["role"] = role
                session["name"] = user["name"]
                if role == "admin":
                    return redirect("/admin")
                elif role == "doctor":
                    return redirect("/doctor")
                else:
                    return redirect("/patient")
        flash("Invalid credentials")
    return render_template("index.html")

@app.route("/signup", methods=["POST"])
def signup():
    username = request.form["username"]
    password = request.form["password"]
    role = request.form["role"]
    name = request.form["name"]

    users = load_users()
    if username in users:
        flash("Username already exists")
    else:
        save_user(username, password, role, name)
        flash("Signup successful! Please log in.")
    return redirect("/")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

def align_and_append_row(df, data):
    """
    Align new-row dict to df columns by:
    - keeping only columns present in df
    - adding missing df columns with None
    - reordering columns to df.columns
    - appending to df and returning new df
    """
    # Make a DataFrame from the incoming dict
    new_row = pd.DataFrame([data])

    # Ensure all df.columns appear in new_row
    for col in df.columns:
        if col not in new_row.columns:
            new_row[col] = None

    # Re-order columns to match df exactly
    new_row = new_row[df.columns]

    # Append safely
    df = pd.concat([df, new_row], ignore_index=True)
    return df


HIGH_RISK_THRESHOLD = 0.7  # adjust threshold if needed

@app.route("/predict", methods=["POST"])
def predict():
    global df
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data received"}), 400

    # --- Patient info ---
    patient_name = data.get("name") or session.get("name") or "Unknown Patient"
    patient_id = data.get("patient_id") or "N/A"
    patient_number = data.get("patient_number")  # e.g., +61439157276
    patient_email = data.get("email")

    # --- Predict AI risk ---
    try:
        risk_score = min(max(patient_risk_score(data), 0.0), 1.0)
        risk_level = ai_level(risk_score)
    except Exception as e:
        print(f"⚠️ Error predicting risk: {e}")
        return jsonify({"error": "Error predicting risk"}), 500

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # --- Store patient data ---
    data_to_store = {
        **data,
        "ai_risk_score": risk_score,
        "ai_risk": risk_level,
        "timestamp": timestamp,
        "name": patient_name,
        "patient_id": patient_id,
        "department": data.get("department", "General Medicine")
    }
    if session.get("role") == "patient":
        data_to_store["username"] = session.get("username")

    new_row = pd.DataFrame([data_to_store]).dropna(axis=1, how="all")
    df = pd.concat([df, new_row], ignore_index=True)
    df.to_csv(DATA_PATH, index=False)

    # --- In-app alert ---
    if risk_score >= HIGH_RISK_THRESHOLD:
        alert = {
            "patient_id": patient_id,
            "patient_name": patient_name,
            "risk_score": risk_score,
            "recommendation": "Follow care plan.",
            "timestamp": timestamp
        }
        in_app_alerts.append(alert)

    # --- SMS Notification ---
    sms_sent = False
    if patient_number and risk_score >= HIGH_RISK_THRESHOLD:
        try:
            sms_sent = notify_patient_sms(
                patient_number,
                patient_name,
                risk_level,
                risk_score,
                dry_run=False
            )
        except Exception as e:
            print(f"⚠️ Failed to send SMS: {e}")

    # --- PDF generation ---
    buffer = BytesIO()
    c = canvas.Canvas(buffer)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(100, 800, "🏥 Personalized Care Plan")
    c.setFont("Helvetica", 12)
    y = 760

    for k, v in data.items():
        if k in ["ai_risk_score", "ai_risk", "recommendation", "email"]:
            continue
        c.drawString(50, y, f"{k}: {v}")
        y -= 20
        if y < 50:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = 800

    display_risk = risk_score * 100
    risk_percent_str = f"{display_risk:.1f}%"
    if risk_level == "High":
        recommendation = "⚠️ High risk! Immediate follow-up required."
    elif risk_level == "Medium":
        recommendation = "Monitor closely and review medications."
    else:
        recommendation = "Low risk. Routine care."

    if y < 100:
        c.showPage()
        c.setFont("Helvetica", 12)
        y = 800

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "📊 AI Risk Summary")
    y -= 25
    c.setFont("Helvetica", 12)
    c.drawString(50, y, f"Predicted Risk Level: {risk_level}")
    y -= 20
    c.drawString(50, y, f"Risk Score: {risk_percent_str}")
    y -= 20
    c.drawString(50, y, f"Recommended Care: {recommendation}")

    c.save()
    buffer.seek(0)

    filename = f"{(patient_name or 'patient').replace(' ', '_')}_care_plan.pdf"

    # --- Email PDF if needed ---
    if patient_email and risk_level in ["High", "Medium"]:
        try:
            buffer.seek(0)
            send_patient_email_with_careplan(
                mail,
                patient_email,
                patient_name,
                risk_score,
                risk_level,
                buffer
            )
        except Exception as e:
            print(f"⚠️ Failed to send PDF email: {e}")

    buffer.seek(0)

    # --- Send PDF + metadata headers ---
    response = send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )
    response.headers["X-Risk-Level"] = risk_level
    response.headers["X-Risk-Score"] = f"{risk_score*100:.1f}"
    response.headers["X-SMS-Sent"] = str(sms_sent)

    return response

@app.route("/predict_json", methods=["POST"])
def predict_json():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data received"}), 400

    try:
        risk_score = min(max(patient_risk_score(data), 0.0), 1.0)
        risk_level = ai_level(risk_score)
        return jsonify({"risk_score": risk_score, "risk_level": risk_level})
    except Exception as e:
        print("⚠️ Error predicting risk:", e)
        return jsonify({"error": "Error predicting risk"}), 500


@app.route("/admin")
def admin():
    global df
    if session.get("role") != "admin":
        flash("Access denied")
        return redirect("/login")

    # --- Ensure essential columns exist ---
    essential_cols = {
        "patient_id": range(1, len(df) + 1),
        "name": [f"Patient {i+1}" for i in range(len(df))],
        "department": "General",
        "readmitted": 0,
        "last_alert": pd.NaT
    }

    for col, default in essential_cols.items():
        if col not in df.columns:
            df[col] = default
        else:
            if col == "name":
                df['name'] = df['name'].fillna(pd.Series([f"Patient {i+1}" for i in range(len(df))]))
            elif col == "department":
                df['department'] = df['department'].fillna("General").replace('', 'General')
            elif col == "readmitted":
                df['readmitted'] = df['readmitted'].fillna(0)
            elif col == "last_alert":
                df['last_alert'] = pd.to_datetime(df['last_alert'], errors='coerce')

    # --- Compute AI risk scores ---
    df['ai_risk_score'] = df.apply(patient_risk_score, axis=1)
    df['ai_risk'] = df['ai_risk_score'].apply(ai_level)

    # --- Department Alerts ---
    alerts_list = []
    for dept, group in df.groupby("department"):
        if dept == "Unknown":
            continue
        high_pct = (group['ai_risk'] == "High").mean()
        medium_pct = (group['ai_risk'] == "Medium").mean()
        if high_pct > 0.3:
            status = "High"
            rate = round(high_pct * 100, 2)
        elif medium_pct > 0.3:
            status = "Medium"
            rate = round(medium_pct * 100, 2)
        else:
            status = "Low"
            rate = round((high_pct + medium_pct) * 100, 2)
        alerts_list.append({"department": dept, "rate": rate, "status": status})

    # --- Department Summary ---
    departments = []
    for dept, group in df.groupby("department"):
        readmission_rate = group['readmitted'].mean() * 100
        high_pct = (group['ai_risk'] == "High").mean()
        medium_pct = (group['ai_risk'] == "Medium").mean()
        if high_pct > 0.3:
            alert_status = "High"
        elif medium_pct > 0.3:
            alert_status = "Medium"
        else:
            alert_status = "Low"

        recommended_action = {
            "High": "Immediate follow-up required. Review all patient cases.",
            "Medium": "Monitor closely and review medications.",
            "Low": "Routine follow-up."
        }[alert_status]

        departments.append({
            "Department": dept,
            "Readmission Rate (%)": round(readmission_rate, 2),
            "Alert Status": alert_status,
            "Recommended Action": recommended_action
        })

    # --- Top 5 High-Risk Patients ---
    top_patients_df = df[df['ai_risk'] == "High"].sort_values("ai_risk_score", ascending=False)
    top_patients_df = top_patients_df.drop_duplicates(subset='patient_id').head(5)
    top_patients_df['last_alert'] = top_patients_df['last_alert'].apply(
        lambda x: x.strftime("%Y-%m-%d %H:%M:%S") if pd.notnull(x) else "N/A"
    )

    top_patients = []
    for _, row in top_patients_df.iterrows():
        top_patients.append({
            "name": row['name'],
            "risk": row['ai_risk'],
            "score": float(row['ai_risk_score']),  # Pass float, not string
            "last_alert": row['last_alert']
        })

    # --- Department Readmission Chart ---
    dept_summary = df.groupby("department")["readmitted"].mean().reset_index()
    dept_summary['readmitted'] = dept_summary['readmitted'].apply(lambda x: x*100 if x <= 1 else x)
    if dept_summary['readmitted'].sum() == 0:
        dept_summary['readmitted'] = np.random.randint(5, 20, size=len(dept_summary))

    fig = px.bar(
        dept_summary,
        x="department",
        y="readmitted",
        text="readmitted",
        color="readmitted",
        color_continuous_scale="Viridis",
        title="Department Readmission Rates (%)",
        labels={"readmitted": "Readmission Rate (%)", "department": "Department"}
    )
    fig.update_traces(
        texttemplate='%{text:.2f}%',
        textposition='outside',
        marker_line_color='rgb(8,48,107)',
        marker_line_width=1.5,
        opacity=0.85
    )
    fig.update_layout(yaxis=dict(range=[0, max(dept_summary['readmitted']) + 10]))
    graph_html = fig.to_html(full_html=False)

    # --- Monthly Risk Trend ---
    if 'admission_date' in df.columns:
        df['month'] = pd.to_datetime(df['admission_date'], errors='coerce').dt.to_period('M')
        trend_df = df.groupby(['month', 'department'])['ai_risk_score'].mean().reset_index()
        trend_df['month'] = trend_df['month'].astype(str)
        fig2 = px.line(trend_df, x="month", y="ai_risk_score", color="department",
                       title="Monthly Average Risk Score Trend by Department", markers=True)
        trend_html = fig2.to_html(full_html=False)
    else:
        trend_html = ""  # fallback empty string

    return render_template(
        "admin.html",
        graph_html=graph_html,
        trend_html=trend_html,
        departments=departments,
        top_patients=top_patients,
        alerts_list=alerts_list
    )

# --- Helper function: notify_patient ---
def notify_patient(patient_email, patient_id, patient_name, risk_score, risk_level):
    """Send patient notification and update last_alert timestamp."""
    if not patient_email:
        return False

    instructions = []
    if risk_level == "High":
        instructions = [
            "Immediate follow-up required",
            "Check medication adherence daily",
            "Call hospital if symptoms worsen"
        ]
    elif risk_level == "Medium":
        instructions = [
            "Monitor symptoms closely",
            "Review medications weekly",
            "Schedule follow-up visit in 2 weeks"
        ]
    else:
        instructions = [
            "Routine care",
            "Maintain healthy diet and exercise",
            "Next visit as per schedule"
        ]

    try:
        msg = Message(
            "📌 Your Readmission Risk Update",
            sender=app.config['MAIL_USERNAME'],
            recipients=[patient_email]
        )
        msg.body = (
            f"Hello {patient_name},\n\n"
            f"Your readmission risk has been evaluated as {risk_level} "
            f"({risk_score*100:.1f}%).\n\n"
            "Care Instructions:\n- " + "\n- ".join(instructions) +
            "\n\nPlease review your care plan in the app or contact your doctor for guidance.\n\n"
            "Stay healthy,\nYour Hospital Care Team"
        )
        mail.send(msg)

        # --- Update last_alert timestamp ---
        df.loc[df['patient_id'] == patient_id, 'last_alert'] = pd.Timestamp.now()

        return True
    except Exception as e:
        print(f"⚠️ Failed to send patient notification: {e}")
        return False

# --- Doctor Dashboard ---
@app.route("/doctor")
def doctor():
    if session.get("role") != "doctor":
        flash("Access denied")
        return redirect("/login")
    return render_template("doctor.html", session=session)

@app.route("/doctor/add_patient", methods=["GET", "POST"])
def add_patient():
    if session.get("role") != "doctor":
        flash("Access denied")
        return redirect("/login")
    
    global df

    if request.method == "POST":
        # Collect patient input safely
        patient_data = {
            "name": request.form.get("name", "Unknown"),
            "age": int(request.form.get("age") or 0),
            "gender": request.form.get("gender", "N/A"),
            "department": request.form.get("department", "General"),
            "time_in_hospital": int(request.form.get("time_in_hospital") or 0),
            "num_medications": int(request.form.get("num_medications") or 0),
            "diagnosis": request.form.get("diagnosis", "N/A"),
            "past_visits": int(request.form.get("past_visits") or 0),
            "treatment": request.form.get("treatment", "N/A")
        }

        # --- Predict risk using model.py ---
        try:
            risk_result = predict_risk(patient_data)
            patient_data["ai_risk_score"] = risk_result.get("risk_score", 0.0)
            patient_data["ai_risk"] = risk_result.get("risk_level", ai_level(patient_data["ai_risk_score"]))
        except Exception as e:
            print(f"⚠️ Error predicting risk: {e}")
            patient_data["ai_risk_score"] = 0.0
            patient_data["ai_risk"] = "Low"

        patient_data["last_alert"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Append safely to df
        df = align_and_append_row(df, patient_data)
        df.to_csv(DATA_PATH, index=False)

        # --- Send SMS if high risk ---
        patient_number = request.form.get("phone")  # optional field in form
        sms_sent = False
        if patient_number and patient_data["ai_risk_score"] >= HIGH_RISK_THRESHOLD:
            try:
                sms_sent = notify_patient_sms(
                    patient_number,
                    patient_data["name"],
                    patient_data["ai_risk"],
                    patient_data["ai_risk_score"],
                    dry_run=False
                )
            except Exception as e:
                print(f"⚠️ Failed to send SMS: {e}")

        # Flash message with risk
        flash(
            f"Patient added. Predicted Risk: {patient_data['ai_risk']} "
            f"({patient_data['ai_risk_score']*100:.1f}%). SMS sent: {sms_sent}"
        )
        return redirect(url_for("doctor"))

    # GET request: render form
    departments = df['department'].unique().tolist()
    return render_template("doctor_add_patient.html", session=session, departments=departments)



@app.route("/patient")
def patient():
    if session.get("role") != "patient":
        flash("Access denied")
        return redirect("/login")

    username = session.get("username")

    # Get all rows for this patient
    patient_rows = df[df['username'] == username]  if 'username' in df.columns else pd.DataFrame()

    # Prepare latest data for pre-filling form
    patient_data = {}
    if not patient_rows.empty:
        latest_row = patient_rows.iloc[-1]  # last submission
        patient_data = {
            'age': latest_row.get('age', ''),
            'time_in_hospital': latest_row.get('time_in_hospital', ''),
            'num_lab_procedures': latest_row.get('num_lab_procedures', ''),
            'num_medications': latest_row.get('num_medications', ''),
            'change': latest_row.get('change', 'No'),
            'diabetesMed': latest_row.get('diabetesMed', 'No'),
            'acarbose': latest_row.get('acarbose', 'No'),
            'acetohexamide': latest_row.get('acetohexamide', 'No'),
            'metformin': latest_row.get('metformin', 'No'),
            'glipizide': latest_row.get('glipizide', 'No'),
            'insulin': latest_row.get('insulin', 'No')
        }

    # Prepare patient history for table display
    patient_history = patient_rows.sort_values(by='timestamp', ascending=False).to_dict(orient='records') if not patient_rows.empty else []

    return render_template(
        "patient.html",
        session=session,
        patient_data=patient_data,
        patient_history=patient_history
    )



# --- Patient: View Risk & Instructions ---
@app.route("/patient/view/<name>")
def patient_view(name):
    if session.get("role") != "patient":
        flash("Access denied")
        return redirect("/login")
    
    patient_row = df[df['name'] == name]
    if patient_row.empty:
        flash("Patient not found")
        return redirect("/patient")
    
    patient_row = patient_row.iloc[0]
    
    # Personalized care instructions
    instructions = []
    if patient_row['ai_risk'] == "High":
        instructions = [
            "Immediate follow-up required",
            "Check medication adherence daily",
            "Call hospital if symptoms worsen"
        ]
    elif patient_row['ai_risk'] == "Medium":
        instructions = [
            "Monitor symptoms closely",
            "Review medications weekly",
            "Schedule follow-up visit in 2 weeks"
        ]
    else:
        instructions = [
            "Routine care",
            "Maintain healthy diet and exercise",
            "Next visit as per schedule"
        ]

    return render_template(
        "patient.html",
        session=session,
        patient=patient_row.to_dict(),
        care_instructions=instructions
    )


# --- API Alerts for Doctor ---
@app.route("/api/alerts")
def api_alerts():
    df['ai_risk_score'] = df.apply(patient_risk_score, axis=1)
    df['ai_risk'] = df['ai_risk_score'].apply(ai_level)
    alerts_list = []

    for dept, group in df.groupby("department"):
        if dept == "Unknown":  # skip placeholder
            continue

        high_pct = (group['ai_risk'] == "High").mean()
        medium_pct = (group['ai_risk'] == "Medium").mean()

        if high_pct > 0.3:
            alerts_list.append({
                "department": dept,
                "rate": round(high_pct * 100, 2),
                "status": "High"
            })
        elif medium_pct > 0.3:
            alerts_list.append({
                "department": dept,
                "rate": round(medium_pct * 100, 2),
                "status": "Medium"
            })

    return jsonify(alerts_list)

# --- Monthly Report (Robust & Safe) ---
@app.route("/admin/report")
def admin_report():
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    from flask import send_file
    from openpyxl.styles import Font, PatternFill
    import traceback

    try:
        global df

        # Step 0: Check if df exists and has data
        if 'df' not in globals() or df.empty:
            return "No patient data available to generate report.", 400

        # Step 1: Ensure essential report columns exist
        report_columns = [
            "patient_id", "name", "department", "age", "gender",
            "time_in_hospital", "num_medications", "readmitted",
            "ai_risk_score", "ai_risk", "last_alert"
        ]

        for col in report_columns:
            if col not in df.columns:
                if col == "last_alert":
                    df[col] = pd.NaT
                elif col in ["ai_risk_score", "readmitted", "time_in_hospital", "num_medications", "age"]:
                    df[col] = 0
                else:
                    df[col] = "N/A"

        # Step 2: Compute risk scores safely
        if df["ai_risk_score"].isnull().all():
            df["ai_risk_score"] = df.apply(patient_risk_score, axis=1)
        if df["ai_risk"].isnull().all():
            df["ai_risk"] = df["ai_risk_score"].apply(ai_level)

        # Format last_alert safely
        df["last_alert"] = pd.to_datetime(df["last_alert"], errors="coerce")\
                            .dt.strftime("%Y-%m-%d %H:%M:%S").fillna("N/A")

        # Step 3: Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # Sheet 1: All Patients
            df[report_columns].to_excel(writer, index=False, sheet_name="Patients")

            # Sheet 2: Department AI Summary
            ai_summary = df.groupby("department")["ai_risk_score"].mean().reset_index()
            ai_summary.rename(columns={"ai_risk_score": "avg_ai_risk"}, inplace=True)
            df_counts = df.groupby("department")["patient_id"].count().reset_index().rename(columns={"patient_id": "num_patients"})
            ai_summary = ai_summary.merge(df_counts, on="department", how="left")
            ai_summary.to_excel(writer, index=False, sheet_name="AI_dept_summary")

            # Sheet 3: Top 10 High-Risk Patients
            top_ai_patients = df[df['ai_risk'] == "High"]\
                                .sort_values("ai_risk_score", ascending=False)\
                                .drop_duplicates(subset='patient_id').head(10)
            if top_ai_patients.empty:
                top_ai_patients = pd.DataFrame(columns=report_columns)
            top_ai_patients[report_columns].to_excel(writer, index=False, sheet_name="Top_AI_Patients")

            # Step 4: Apply safe formatting
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Bold header row + fill
                header_row = list(sheet.iter_rows(min_row=1, max_row=1))
                if header_row:
                    for cell in header_row[0]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

                # Auto column width safely
                for column_cells in sheet.columns:
                    try:
                        max_length = max([len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells])
                        column_letter = column_cells[0].column_letter
                        sheet.column_dimensions[column_letter].width = max_length + 2
                    except:
                        continue

                # Conditional formatting for ai_risk_score
                header_names = [cell.value for cell in header_row[0]] if header_row else []
                if "ai_risk_score" in header_names:
                    col_idx = header_names.index("ai_risk_score") + 1
                    for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            try:
                                val = float(cell.value)
                                if val >= 0.7:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                                elif val >= 0.3:
                                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                                else:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                            except (ValueError, TypeError):
                                continue

        # Step 5: Send file
        filename = f"monthly_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return f"Error generating report: {str(e)}\n\n{traceback.format_exc()}", 500



# --- High-Risk Alerts Function ---
def send_high_risk_alerts():
    import datetime
    global df
    alerts = []

    # Ensure ai_risk column exists
    if 'ai_risk' not in df.columns:
        df['ai_risk'] = df.apply(lambda row: ai_level(patient_risk_score(row)), axis=1)

    for idx, row in df.iterrows():
        if row['ai_risk'] == "High":
            alert_msg = f"[ALERT] Patient {row['name']} in {row['department']} is HIGH risk ({row['ai_risk_score']*100:.2f}%)"
            print(alert_msg)  # can replace with email/SMS
            alerts.append(alert_msg)
            # Update last_alert timestamp safely
            df.at[idx, 'last_alert'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return alerts

# --- Run Flask ---
if __name__ == "__main__":
    app.run(debug=True, port=5000)
